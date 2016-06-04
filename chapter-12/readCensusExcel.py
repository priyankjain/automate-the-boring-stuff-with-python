#! python3
# readCensusExcel.py - Tabulates population and number of census
# tracts for each county
# Author: Priyank Jain
#####################################################################
import openpyxl, pprint
print("Opening workbook...")
wb = openpyxl.load_workbook('censuspopdata.xlsx')
sheet = wb.get_sheet_by_name('Population by Census Tract')
countyData = {}
print('Reading rows...')
for row in range(2, sheet.max_row + 1):
	state = sheet.cell(row=row, column=2).value
	county = sheet.cell(row=row, column=3).value
	pop = sheet.cell(row=row, column=4).value
	countyData.setdefault(state,{})
	countyData[state].setdefault(county,{'tract':0, 'pop':0})
	countyData[state][county]['tract'] += 1
	countyData[state][county]['pop'] += int(pop)
print('Successfully read the rows')
print('Writing the results...')
wb.save('censuspopdata.xlsx')
resultFile = open('census2010.py', 'w')
resultFile.write('allData = ' + pprint.pformat(countyData))
resultFile.close()
print('Done.')
