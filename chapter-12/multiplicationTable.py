#! python3
# multiplicationTable.py - Creates an NxN multiplication table in an
# Excel spreadsheet
# Author: Priyank Jain
#####################################################################
import openpyxl, sys
if len(sys.argv)!=2:
	print('Usage: python3 multiplicationTable.py <int>')
	sys.exit(-1)
number = None
try:
	number = abs(int(sys.argv[1]))
except ValueError as e:
	print('First argument should be an integer')
	sys.exit(-1)
wb = openpyxl.Workbook()
sheet = wb.active
boldFont = openpyxl.styles.Font(bold=True)
for rowNum in range(1, number+2):
	for colNum in range(1, number+2):
		if rowNum==1 and colNum==1:
			sheet.cell(row=rowNum, column=colNum).value=''
		elif rowNum==1:
			sheet.cell(row=rowNum, column=colNum).value = colNum-1
			sheet.cell(row=rowNum, column=colNum).font = boldFont
		elif colNum==1:
			sheet.cell(row=rowNum, column=colNum).value=rowNum-1
			sheet.cell(row=rowNum, column=colNum).font = boldFont
		else:
			sheet.cell(row=rowNum, column=colNum).value = (rowNum-1)*(colNum-1)
wb.save('table_{0}.xlsx'.format(number))
sys.exit(0)
