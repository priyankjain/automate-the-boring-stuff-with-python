#! python3
# sheetInverter.py - Transposes the rows and columns in a spreadsheet
# Author: Priyank Jain
#####################################################################
import openpyxl, sys, os
if len(sys.argv)>2:
	print('Usage: python3 sheetInverter.py <worbook-name>')
	sys.exit(-1)
if not os.path.exists(sys.argv[1]):
	print('The specified file {0} does not exist'.format(sys.argv[1]))
	sys.exit(-2)
inputFile = sys.argv[1]
outputFile = '{0}_inverted{1}'.format(*os.path.splitext(inputFile))
if os.path.exists(outputFile):
	print('Destination file {0} already exists'.format(outputFile))
	sys.exit(-3)
inputWb = openpyxl.load_workbook(inputFile)
inputSheet = inputWb.active
outputWb = openpyxl.Workbook()
outputSheet = outputWb.active
maxRow = inputSheet.get_highest_row()
maxCol = inputSheet.get_highest_column()
for i in range(1, max(maxRow, maxCol)+1):
	for j in range(1, min(maxRow, maxCol)+1):
		outputSheet.cell(row=i, column=j).value = inputSheet.cell(row=j, column=i).value
		outputSheet.cell(row=j, column=i).value = inputSheet.cell(row=i, column=j).value
outputWb.save(outputFile)
inputWb.save(inputFile)
print('File inverted successfully')
