#! python3
# updateProduceSales.py - Update prices of Garlic, Celery and Lemons
# in an excel sheet of products
# Author: Priyank Jain
#####################################################################
import openpyxl
updatedPrices = {'Garlic': 3.07,
		'Celery': 1.19,
		'Lemon': 1.27}
wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.get_sheet_by_name('Sheet')

for row in range(2, sheet.max_row + 1):
	if(sheet.cell(row=row, column=1).value in updatedPrices):
		sheet.cell(row=row, column=2).value = updatedPrices[sheet.cell(row=row, column=1).value]
		print('Update price in row {0}'.format(row))
wb.save('produceSales_updated.xlsx')
